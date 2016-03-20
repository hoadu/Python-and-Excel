from openpyxl import load_workbook, Workbook
import openpyxl
import os
import itertools


# Display functions
def display_functions(sheet):
    while True:
        print('Functions:')
        print('1 : Get total of a range of cells')
        print('2 : Find best choice order')
        print('3 : Exit')
        choice = input('Your choice: ')

        if '2' in choice:

            # Best cost
            best_cost = 100000000
            # Numbers
            numbers = []
            # Combos
            combinations = []
            # Best combo
            best_combo = ()

            for i in range(2, 16):
                numbers.append(i)

            for s in range(0, len(numbers) + 1):
                for sub in itertools.combinations(numbers, s):
                    combinations.append(sub)

            for k in combinations:
                total_cost = 0
                progress = 0
                for l in k:
                    l = str(l)
                    total_cost += sheet['B' + l].value
                    progress += (sheet['C' + l].value * 100)

                    # If progress is over 100, break.
                    if progress > 100:
                        break

                # If the current combination is more efficient, save it.
                if total_cost < best_cost and progress >= 100.0:
                    best_cost = total_cost
                    best_combo = k

            for g in best_combo:
                g = str(g)
                print(g + ' : {0}'.format(sheet['B' + g].value) + ' | {0}'.format(sheet['C' + g].value * 100) + '%')

        # Make a new line
        print('')

        if '3' in choice:
            clear()
            break


def display_manipulations(workb):
    while True:
        print('Manipulations:')
        print('1 : View a cells contents')
        choice = input('Your choice: ')

        # Display a cells contents
        if '1' in choice:
            sheet_ranges = workb['range names']
            cell = input('Which cell would you like to view: ')
            print(sheet_ranges[cell].value)


def clear():
    os.system('cls')


# Variables
user_documents = os.path.expanduser('~\Documents')  # Users documents dir.
save_directory = user_documents + '\Excel in Python'  # Save the default save dir.
sheets = []


# Check if the default save location exists
if not os.path.exists(save_directory):
    # Make the directory
    os.mkdir(save_directory)
    # Inform the user
    print('[INFO] Program save directory created: ' + save_directory)

# Check if the workbook save directory exists.
if not os.path.exists(save_directory + '\Saved Workbook Directories.txt'):
    # Create a txt file to hold save directories
    saved_workbooks = open(save_directory + '\Saved Workbook Directories.txt', 'a')
    # Inform the user
    print('[INFO] Workbook save directory created: ' + save_directory + '\Saved Workbook Directories.txt')

# Prompt the user on whether or not to use an existing workbook.
use_existing = input('Would you like to use an existing workbook (y/n): ')

# Select the workbook
while True:
    try:
        if 'y' in use_existing:
            # Get the workbook directory/name
            workbook_directory = input('Enter the directory where the workbook is stored: ')

            try:
                # Load the workbook.
                wb = load_workbook(workbook_directory)
                # Notify the user the workbook has been loaded successfully.
                ws = wb.active
                print('[INFO] Successfully loaded.')
            except openpyxl.utils.exceptions.InvalidFileException:
                # Get the filename and file_extension of the given file.
                filename, file_extension = os.path.splitext(workbook_directory)
                # Print the issue
                print('[ERROR] Incorrect file type \'' + file_extension + '\'.')
                # Continue
                continue
        else:
            # Prompt the user for the filename.
            filename = input('Enter the name of the new workbook: ')
            # Create the new workbook
            wb = Workbook(filename)
            # Save the workbook
            ########

    except FileNotFoundError:
        print('[INFO] File \"' + workbook_directory + '\" not found.')
        continue
    break

# Get all the sheet names
for sheet in wb.get_sheet_names():
    sheets.append(sheet)

while True:
    print('\nPossible actions:')
    print('1 : Manipulate the workbook')
    print('2 : Functions')
    print('4 : Change the current workbook')
    print('3 : Exit')
    choice = input('Your choice: ')

    # Display possible manipulations
    if '1' in choice:
        clear()
        display_manipulations(wb)
        continue

    # Display all functions
    if '2' in choice:
        clear()
        display_functions(ws)
        continue

    # Exit the program
    if '3' in choice:
        print('\n[INFO] Exiting...')
        break
